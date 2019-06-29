
# coding: utf-8

# In[ ]:


import pandas as pd
import numpy as np
from math import isnan
from itertools import zip_longest
import os
import re

if not os.path.exists("Labelling"):
    os.mkdir("Labelling") 


# In[ ]:


Groups_File_Name = str(input("Please give the name of the file with groups.\n"))


# In[ ]:


Groups_df = pd.read_excel(Groups_File_Name, dtype=str)


# In[ ]:


Groups_dict = Groups_df.to_dict("list")

#cleaning the list from nans
for Group in Groups_dict:
    Groups_dict[Group] = [str(word) for word in Groups_dict[Group] if str(word) != 'nan']
    print(Groups_dict[Group])


# In[ ]:


Regex_dict = {}


# In[ ]:


for Group_Key in Groups_dict.keys():
    print(Group_Key)
    Regex_dict[Group_Key] = "(" + "|".join(Groups_dict[Group_Key]) + ")"
Regex_dict


# In[ ]:


Label_File_name = str(input("Please give the name of the FILE with lines to be labelled.\n"))
Label_col = str(input("Please give the name of the COLUMN with lines to be labelled.\n"))


# In[ ]:


Label_df = pd.read_excel(Label_File_name)

for Group_Key in Regex_dict:
    Label_df.loc[:, Group_Key] = Label_df[Label_col].str.findall(Regex_dict[Group_Key], re.IGNORECASE)
Label_df


# In[ ]:


from datetime import datetime
timestamp= datetime.now().strftime("%Y%m%d-%H%M%S")
Label_df.to_excel("Labelling\\Label_"+timestamp+"_"+Label_File_name)


# In[ ]:


"""

Без кавычек:

#%load file.py
%%writefile file.py  - в начале блока
%pycat  -
%run file.py
%lsmagic

from IPython.core.interactiveshell import InteractiveShell
InteractiveShell.ast_node_interactivity = "all"

===openpyxl===
from openpyxl import load_workbook
from openpyxl import Workbook

wb = load_workbook()
wb_ws = wb.get_active_sheet()

wrwb = Workbook()
wrwb_ws = wrwb.get_active_sheet()

wb.save()

===numpy/pandas===
import pandas as pd
import numpy as np

excel_df = pd.read_excel()
csv_df = pd.read_csv()


df.to_excel()
df.to_csv()

writer = pd.ExcelWriter('',engine='xlsxwriter',options={})
df.to_excel(writer)
writer.save()


====requests/BeautifulSoup===
import requests
from bs4 import BeautifulSoup

page = requests.get("http://yandex.ru")
page.encoding = "windows-1251"
soup = BeautifulSoup(''.join(page.text), "html.parser\"),
soup.findAll("div")


===Files and directories===
import os
FileList = os.listdir()

#if not os.path.exists("Folder"):
#   os.mkdir("Folder") 

"""

